from fpdf import FPDF
from module1 import PDF_MC_Table
import random
import pandas as pd
import numpy as np
import openpyxl
import datetime
print("hello")
fontSize=10
height=7

#Read an excel file that contains 2 workbooks named agenda an roleplayers
# the function returns only one dataframe
def get_final_df(xlsFile):
    xls = pd.ExcelFile(xlsFile)
    df1 = pd.read_excel(xls, 'agenda')
    df2 = pd.read_excel(xls, 'roleplayers')
    print(f"printing agenda workbook {df1}")
    print(f"printing role players workbook {df2}")

    rolePlayers=dict()

    for index, row in df2.iterrows():
        rolePlayers[row[0]]=row[1] #dictionary from df, used this function because wanted in a particular format
    print(f" printing role player dictionary{rolePlayers}")
    for i in range(0, len(df1)):
        key=df1.loc[i,'role players'] #finding the role player written in agenda
        #iterating over the dictionary keys to find the rol player, if present
        #subsitute the value of the role player
        if key in rolePlayers.keys():
            print("PRINTING ROLE AND ROLE PLAYER NAME")
            print(key,rolePlayers[key])
            df1.loc[i,'role players']=rolePlayers[key]
    print("inside get final df function that returns the role player name df")
    print(df1)
    return df1  # return the final dataframe with role player names for agenda

# this function will take the ending time of the previous activity and add the allotted time to it
# returns the new time for next activity
def new_time(time_str,time_allotted):
    time_calc=[]
    if "nan" not in time_str and "nan" not in time_allotted:

        time_str=time_str.split(" pm")[0]
        date = datetime.datetime.strptime(time_str, '%H:%M') #print(date.strftime('%H:%M:%S'))
        if "-" in time_allotted:
            
            time_allotted=time_allotted.replace("mins","")
            time_mins=int(time_allotted.split("-")[1].lstrip())
            time_secs=time_mins*60
            #print("hyphen time -----",time_secs)
            #time_str = '7:55'
            date = date+datetime.timedelta(seconds=time_secs) #print(date)  # 👉️ 2023-09-24 09:30:
            x=date.strftime('%H:%M')
            #date = date+datetime.timedelta(seconds=3600)
            time_calc=str(x) + " pm" 
            return time_calc
            #print("when only minutes time and hyphen",time_calc)

        elif "sec" in time_allotted: # 1 min 30 sec
            time=time_allotted.split(" min ") # 1,30 sec 
            time_min=int(time[0]) #1
            time_sec=int(str(time[1]).split("sec")[0]) #30
            #print("COMBINED TIME IS")  
            #print("min sec time is -----",time_min,time_sec)  
            time_to_add=(time_min*60)+(60)
            #date = datetime.datetime.strptime(time_str, '%H:%M') #print(date.strftime('%H:%M:%S'))
            date = date+datetime.timedelta(seconds=time_to_add) #print(date)  # 👉️ 2023-09-24 09:30:
            x=date.strftime('%H:%M')
            time_calc=str(x) + " pm"
            return time_calc 
            #print(f"when both min and sec is present in the string {time_calc}")
        
        else:
            time=int(time_allotted.split(" min")[0])
            time_to_add=time*60
            date = date+datetime.timedelta(seconds=time_to_add) #print(date)  # 👉️ 2023-09-24 09:30:
            x=date.strftime('%H:%M')
            time_calc=str(x) + " pm" 
            return time_calc
            #print(f"when only minute is given in the time {time_calc}")




    elif "nan" in time_str and "nan" in time_allotted:
        return time_calc
        #print(f"when neither time not time allotted {time_calc}") 

    elif "nan" not in time_str and "nan" in time_allotted:
        time_calc=time_str
        return time_calc
        #print(f"when time present but not time allotted{time_calc}") 

    #print("-----FINAL TIME IS :",time_calc)
    
    print("\n")

#fetches the new time from new_time function and edits it in the df
def timing(df):
    df_exp=df.copy()
    #df_exp=df_exp[['time', 'time allotted']]
    df_exp['time']='6:30 pm'
    #print(df_exp)
    for i in range(0, len(df_exp)):
        if i!=0:
            time = str(df_exp.loc[i-1, 'time']).lower()
            time_allotted=str(df_exp.loc[i-1, 'time allotted']).lower()
            #print(f"---- original time is: {time} and time to add is : {time_allotted}")
            df_exp.loc[i,'time']=str(new_time(time,time_allotted))
            #print(df_exp.loc[i,'time'])
    #print(df_exp)
    df_exp=df_exp.replace(np.nan, '')
    return df_exp

#makes agenda pdf  
def make_pdf(df):
    pdf = PDF_MC_Table()
    pdf.add_page()
    pdf.set_font(family='Times', style='B', size=fontSize)
    pdf.set_text_color(256, 256, 256) #white
    # Table with 20 rows and 4 columns
    pdf.SetWidths([20, 110, 35, 25])
    pdf.set_font(family='Times',size=12,style='B')

    #pdf.set_draw_color(0,80,180) # gves colors to the border of the cell
    pdf.set_fill_color(21, 107, 136) # blue
    pdf.cell(w=0, h=height, txt="Connected Across Miles Toastmasters Club", align='C', ln=1,border=1,fill=True)

    #pdf.set_font(family='Times',size=14)
    pdf.set_fill_color(245, 243, 244) #grey
    pdf.set_text_color(142, 54, 78) #red
    pdf.cell(w=0, h=height, txt="Division K | Area 03 | District 124", align='C', ln=1,border=1,fill=True)

    #pdf.set_font(family='Times',size=14)
    pdf.set_fill_color(251, 249, 197) #yellow
    pdf.set_text_color(21, 107, 136) # blue
    pdf.cell(w=0, h=height, txt="Meeting No. 2 | Agenda - 2nd July 2023 | 6:30 PM",align='C', ln=1,border=1,fill=True)
    pdf.set_text_color(0,0,0)
    pdf.set_font(family='Times', style='B', size=fontSize)
    pdf.set_fill_color(142, 54, 78) #red
    pdf.set_text_color(256, 256, 256) #white
    pdf.cell(w=20, h=height, txt="Time",border=1,fill=True)
    pdf.cell(w=110, h=height, txt="Agenda",border=1,fill=True)
    pdf.cell(w=35, h=height, txt="Role Player", border=1,fill=True)
    pdf.cell(w=25, h=height, txt="Time allotted",border=1,ln=1,fill=True)

    #df=get_final_df('Automate-everything-with-python/AgendaRolePlayerAutomation/CAMTMagenda.xlsx')
    
    df_time=timing(df)
    pdf.set_font(family='Times', size=fontSize)
    pdf.set_text_color(0, 0, 0) #black
    for index, row in df_time.iterrows():
        pdf.Row([str(row[0]), str(row[1]), str(row[2]), str(row[3])])

    #end of the meeting PO
    pdf.set_text_color(256, 256, 256) #white
    pdf.set_fill_color(142, 54, 78)
    pdf.cell(w=20, h=height, txt="8:00 PM",border=1,fill=True)
    pdf.set_font(family='Times', style='B', size=fontSize)
    pdf.cell(w=170, h=height, txt="President Adjourns the meeting", align='C', border=1,ln=1,fill=True)
    pdf.set_fill_color(21, 107, 136)
    pdf.cell(w=0, h=height, txt="Networking Session", align='C', border=1,fill=True)
    #PRINTING THE PDF
    #pdf.output("Automate-everything-with-python/AgendaRolePlayerAutomation/OutPut/MeetingNo2.pdf")
    return df_time,pdf


'''
#EXPERIMENT WITH THE ROWS
#df_time=df_time.iloc[:,[2,1,0]] # re arrange the columns 
#df_time = df_time.iloc[:-1] # remove the last row
 df=df.replace(np.nan, '') # replace nan values in df
df = df.iloc[: , :-1] # remove the last column
df = df.iloc[:-1] # remove the last row
df=df.iloc[:,[2,1,0,3]] # re arrange the columns 

#reading from a csv file
def dataframe(csv):
    df = pd.read_csv(csv)
   
    df['agenda'] = df['agenda'].str.wrap(100)
    #df.rename(columns = {"Time Taken":'Time', 'Role':'Agenda','Name':'Role Player','Time Range':'Time Allotted'}, inplace = True)
    #print(df)
    return df

'''